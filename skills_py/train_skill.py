# skills_py/train_skill.py
import os
import json
import re

from brain_state import COMMANDS
import memory
from memory import conn, save_skill

import openpyxl
from openpyxl import Workbook, load_workbook

# Path for your master skill Excel file
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "skills_master.xlsx")

triggers = ["teach", "learn", "add skill"]

def append_skill_to_excel(trigger, action, command, path=EXCEL_PATH):
    # If file exists, load it; otherwise create with headers
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Trigger Phrase", "Action", "Command"])
    # Append the new skill row
    ws.append([trigger, action, command])
    wb.save(path)

def run(user_input):
    # Parse: teach 'trigger' runs 'command'
    match = re.search(r"'(.+?)'\s+runs\s+'(.+?)'", user_input)
    if not match:
        print("[⚠️] Format: teach 'trigger phrase' runs 'command'")
        return {"action": "system", "path_or_command": "echo Invalid format"}

    trigger_phrase, command = match.groups()
    trigger = trigger_phrase.lower()
    action = "system"

    # 1) Write JSON skill file
    filename = trigger.replace(" ", "_") + ".json"
    skill_path = os.path.join("skills", filename)
    new_skill = {
        "triggers": [trigger_phrase],
        "action": action,
        "path_or_command": command
    }
    try:
        with open(skill_path, "w") as f:
            json.dump(new_skill, f, indent=2)
    except Exception as e:
        print(f"[❌] Failed saving JSON: {e}")
        return {"action": "system", "path_or_command": "echo Save error"}

    # 2) Update in-memory commands
    COMMANDS[trigger] = {"action": action, "path_or_command": command}

    # 3) Persist to SQLite
    save_skill(conn, trigger, action, command)

    # 4) Append to Excel
    try:
        append_skill_to_excel(trigger, action, command)
    except Exception as e:
        print(f"[⚠️] Excel write error: {e}")

    # 5) Confirm to user
    print(f"[✅] Learned new skill: '{trigger_phrase}' → {command}")
    return {"action": "system", "path_or_command": f"echo Learned: {trigger_phrase}"}
